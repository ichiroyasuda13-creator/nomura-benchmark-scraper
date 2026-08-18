"""Stage 6: Multi-Sheet Styled Excel and CSV Report Generation.

Generates:
  1. Sheet 1「ファンド一覧」: All funds with AUM, Net Flow, Theme, Benchmark, Distributors.
  2. Sheet 2「MSCI営業ターゲット」: High net inflow non-MSCI funds with action triggers (Who to Call).
  3. Sheet 3「販売会社別ファンド一覧」: Grouped rankings per distributor (matching industry magazine format).
  4. Sheet 4「販社×テーマ別売れ行き」: Broker sales matrix showing which broker sells which theme best.
  5. Sheet 5「商品企画・組成提案」: Consultative proposal cards for asset manager pitches.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from loguru import logger

from app.config import BENCHMARKS_JSON, OUTPUT_DIR
from app.distributors import (
    build_broker_theme_sales_matrix,
    get_funds_grouped_by_distributor,
)
from app.http_client import load_json
from app.models import BenchmarkRecord, format_aum_oku, format_inflow_oku
from app.proposal_generator import generate_product_proposals

OUTPUT_COLUMNS = [
    "rank",
    "management_company",
    "fund_name",
    "fund_code",
    "aum",
    "estimated_net_inflow",
    "performance_effect",
    "aum_date",
    "theme_category",
    "is_etf",
    "fund_type",
    "benchmark",
    "benchmark_detail",
    "index_provider",
    "is_msci",
    "top_distributors",
    "primary_broker",
    "sales_pitch_action",
    "reference_index",
    "confidence",
    "extraction_method",
    "prospectus_pdf_url",
    "source_page_detail_url",
    "note",
    "needs_review",
    "manual_override",
    "review_comment",
]


def records_to_dataframe(records: list[BenchmarkRecord]) -> pd.DataFrame:
    rows = [record.model_dump(mode="json") for record in records]
    frame = pd.DataFrame(rows)
    for column in OUTPUT_COLUMNS:
        if column not in frame.columns:
            frame[column] = None
    return frame[OUTPUT_COLUMNS]


def _style_header(ws: openpyxl.worksheet.worksheet.Worksheet, row_num: int, max_col: int, fill_color: str = "1E293B") -> None:
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    ws.row_dimensions[row_num].height = 26


def _auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            length = sum(1.7 if ord(c) > 127 else 1.0 for c in val_str)
            if length > max_len:
                max_len = int(length)
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10)


def create_styled_excel(records: list[BenchmarkRecord], filepath: Path) -> None:
    wb = openpyxl.Workbook()

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    msci_yes_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    msci_yes_font = Font(name="Meiryo UI", size=9, color="065F46", bold=True)
    target_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    target_font = Font(name="Meiryo UI", size=9, color="991B1B", bold=True)
    regular_font = Font(name="Meiryo UI", size=9)

    company_name = records[0].management_company if records else "野村アセットマネジメント"

    # ── Sheet 1: All Funds (ファンド一覧) ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "ファンド一覧"

    headers1 = [
        "順位", "運用会社", "ファンド名", "コード", "純資産 (億円)", "推定純流入 (億円)", "運用効果 (億円)",
        "テーマ分類", "ベンチマーク指数", "指数提供者", "MSCI採用", "主要販売会社 (Broker)", "営業ターゲット判定",
        "信頼度", "交付目論見書PDF",
    ]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1), fill_color="1E293B")

    for row_idx, r in enumerate(records, start=2):
        oku_aum = round(r.aum / 1e8, 1) if r.aum else 0.0
        oku_inflow = round(r.estimated_net_inflow / 1e8, 1) if r.estimated_net_inflow else 0.0
        oku_perf = round(r.performance_effect / 1e8, 1) if r.performance_effect else 0.0

        row_data = [
            r.rank,
            r.management_company,
            r.fund_name,
            r.fund_code,
            oku_aum,
            oku_inflow,
            oku_perf,
            r.theme_category,
            r.benchmark or "—",
            r.index_provider,
            "はい" if r.is_msci else "いいえ",
            r.top_distributors or r.primary_broker or "主要証券",
            r.sales_pitch_action or ("既存採用（防衛）" if r.is_msci else "🎯 提案対象"),
            r.confidence.value if hasattr(r.confidence, "value") else str(r.confidence),
            r.prospectus_pdf_url or "",
        ]
        ws1.append(row_data)

        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in (5, 6, 7):
                cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (1, 2, 4, 8, 10, 11, 14):
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 11 and r.is_msci:
                cell.fill = msci_yes_fill
                cell.font = msci_yes_font
            elif col_idx == 13 and not r.is_msci and (r.estimated_net_inflow > 0 or r.aum > 1e11):
                cell.fill = target_fill
                cell.font = target_font

    _auto_fit_columns(ws1)

    # ── Sheet 2: MSCI Sales Targets (MSCI営業ターゲット) ─────────────────
    ws2 = wb.create_sheet(title="MSCI営業ターゲット")
    headers2 = [
        "順位", "運用会社", "ファンド名", "コード", "純資産 (億円)", "推定純流入 (億円)", "テーマ分類",
        "現在のベンチマーク", "現在の指数提供者", "主要販売会社 (攻めどころ)", "営業提案アクション (Who to Call)",
    ]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2), fill_color="0F766E")

    non_msci_records = [r for r in records if not r.is_msci and r.aum > 0]
    non_msci_records.sort(key=lambda x: (x.estimated_net_inflow, x.aum), reverse=True)

    for row_idx, r in enumerate(non_msci_records, start=2):
        oku_aum = round(r.aum / 1e8, 1)
        oku_inflow = round(r.estimated_net_inflow / 1e8, 1) if r.estimated_net_inflow else 0.0

        row_data = [
            r.rank,
            r.management_company,
            r.fund_name,
            r.fund_code,
            oku_aum,
            oku_inflow,
            r.theme_category,
            r.benchmark or "—",
            r.index_provider,
            r.top_distributors or r.primary_broker or "主要証券",
            r.sales_pitch_action or f"🎯 {r.primary_broker or '主要販社'} 経由でMSCIリプレイス商談",
        ]
        ws2.append(row_data)
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in (5, 6):
                cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                cell.alignment = Alignment(horizontal="right")
            if col_idx == 11:
                cell.font = Font(name="Meiryo UI", size=9, bold=True, color="0F766E")

    _auto_fit_columns(ws2)

    # ── Sheet 3: Distributor Rankings (販売会社別ファンド一覧) ────────────
    ws3 = wb.create_sheet(title="販売会社別ファンド一覧")
    current_row = 1

    dist_groups = get_funds_grouped_by_distributor(records)
    for dist_name, funds in dist_groups.items():
        # Section Title Bar
        sec_cell = ws3.cell(row=current_row, column=1, value=f"🏛️ {dist_name} 取扱上位ファンド（残高順）")
        sec_cell.font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
        sec_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws3.row_dimensions[current_row].height = 24
        current_row += 1

        # Table Header
        dist_headers = ["順位", "運用商品名 (ファンド名)", "運用会社", "残高 (億円)", "推定純流入 (億円)", "ベンチマーク指数", "MSCI採用", "提案戦略"]
        for c_idx, h in enumerate(dist_headers, start=1):
            c = ws3.cell(row=current_row, column=c_idx, value=h)
            c.font = Font(name="Meiryo UI", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[current_row].height = 20
        current_row += 1

        # Data Rows
        for f in funds:
            r_data = [
                f["rank"],
                f["fund_name"],
                f["management_company"],
                f["aum_oku"],
                f["inflow_oku"],
                f["benchmark"],
                "はい" if f["is_msci"] else "いいえ",
                f["action"],
            ]
            for c_idx, val in enumerate(r_data, start=1):
                cell = ws3.cell(row=current_row, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                if c_idx in (4, 5):
                    cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                    cell.alignment = Alignment(horizontal="right")
                elif c_idx in (1, 3, 7):
                    cell.alignment = Alignment(horizontal="center")
                if c_idx == 7 and f["is_msci"]:
                    cell.fill = msci_yes_fill
                    cell.font = msci_yes_font
                elif c_idx == 7 and not f["is_msci"]:
                    cell.fill = target_fill
                    cell.font = target_font
            current_row += 1

        # Blank line between distributor blocks
        current_row += 1

    _auto_fit_columns(ws3)

    # ── Sheet 4: Broker x Theme Sales Matrix (販社×テーマ別売れ行き) ────────
    ws4 = wb.create_sheet(title="販社×テーマ別売れ行き")
    headers4 = [
        "主要販売会社 (Broker)", "得意テーマ分類", "取扱本数", "純資産総額 (億円)", "推定純流入 (億円)", "商品企画・組成推奨戦略",
    ]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4), fill_color="4338CA")

    matrix_rows = build_broker_theme_sales_matrix(records)
    for row_idx, item in enumerate(matrix_rows, start=2):
        oku_aum = round(item["total_aum"] / 1e8, 1)
        oku_inflow = round(item["total_inflow"] / 1e8, 1)
        pitch = f"{item['broker']} が強い【{item['theme']}】カテゴリーにMSCI指数連動ファンドを新規企画提案"
        row_data = [
            item["broker"],
            item["theme"],
            item["fund_count"],
            oku_aum,
            oku_inflow,
            pitch,
        ]
        ws4.append(row_data)
        for col_idx in range(1, len(headers4) + 1):
            cell = ws4.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in (4, 5):
                cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws4)

    # ── Sheet 5: Consultative Product Proposals (商品企画・組成提案) ─────────
    ws5 = wb.create_sheet(title="商品企画・組成提案")
    headers5 = [
        "提案テーマ", "自社ラインアップ状況", "提案優先度", "推奨MSCI指数", "最適な主幹販社", "市場背景・提案根拠", "具体的なアクションプラン",
    ]
    ws5.append(headers5)
    _style_header(ws5, 1, len(headers5), fill_color="B45309")

    proposals = generate_product_proposals(records, company_name)
    for row_idx, prop in enumerate(proposals, start=2):
        row_data = [
            prop["theme"],
            prop["status"],
            prop["priority"],
            prop["recommended_msci_index"],
            prop["best_selling_brokers"],
            prop["proposal_narrative"],
            prop["action_plan"],
        ]
        ws5.append(row_data)
        for col_idx in range(1, len(headers5) + 1):
            cell = ws5.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in (1, 2, 3):
                cell.alignment = Alignment(horizontal="center")
            if "最優先" in str(prop["priority"]) and col_idx == 3:
                cell.fill = target_fill
                cell.font = target_font

    _auto_fit_columns(ws5)

    wb.save(filepath)


def run_stage6(records: list[BenchmarkRecord] | None = None) -> tuple[Path, Path]:
    if records is None:
        raw = load_json(BENCHMARKS_JSON, [])
        records = [BenchmarkRecord.model_validate(item) for item in raw]
    if not records:
        raise RuntimeError("Stage6 requires benchmark records. Run stage5 first.")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    frame = records_to_dataframe(records)
    csv_path = OUTPUT_DIR / "nomura_benchmarks.csv"
    xlsx_path = OUTPUT_DIR / "nomura_benchmarks.xlsx"

    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")
    create_styled_excel(records, xlsx_path)
    logger.info("Stage6: wrote {} and 5-sheet styled {}", csv_path, xlsx_path)
    return csv_path, xlsx_path
